from fBook import Book
from fShelf import Shelf
def populate(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    s1= wb['Sheet1']
    wb.active(0)
    for i in range(2,s1.max_row+1):
        x=Book(s1.cell(row=i,column='A'),s1.cell(row=i,column='B'), s1.cell(row=i,column='C'))
        y=Shelf.add_book(y,x)
